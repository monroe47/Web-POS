from datetime import timedelta, date
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.conf import settings
from django.core.cache import cache
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from io import BytesIO
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
from POS.utils import get_daily_sales_df
from POS.models import SaleItemUnit, DailySalesRecord, Transaction
from Inventory.models import Item
from .models import ForecastRun, ForecastResult





class SalesForecastDashboardView(LoginRequiredMixin, TemplateView):
    login_url = 'account_management:login'
    """
    Renders the dashboard template and prepares helpful context so the
    front-end can render immediately without extra API calls for static parts.

    Improvements made:
    - Reads optional GET params (horizon, mode) so the view responds to the
      dashboard controls (useful for initial page load).
    - Provides horizon_options, product list, recent sales, top products,
      model_info, latest_run and a small KPI dict so the template can show
      something while the JS fetches the live forecast.
    - Uses defensive try/except blocks so missing POS/Inventory apps or DB
      state won't crash the page — errors fall back to empty lists/defaults.
    - Uses Django cache (short TTL) for heavier aggregates to avoid
      repeated DB load when many users open the dashboard.
    """

    template_name = "Sales_forecast/sales_dashboard.html"
    cache_ttl = 30  # seconds for aggregated blocks; adjust as needed

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Get UI inputs (allow GET override)
        horizon = int(self.request.GET.get("horizon", kwargs.get("default_horizon", 7)))
        mode = self.request.GET.get("mode", kwargs.get("mode", "daily"))
        # Demo mode is disabled by default. Enable only via Django setting
        demo_mode = getattr(settings, 'SALES_FORECAST_ENABLE_DEMO', False)

        ctx["default_horizon"] = horizon
        ctx["mode"] = mode
        ctx["horizon_options"] = [1, 3, 7, 14, 30]
        ctx["demo_mode"] = demo_mode  # ✅ Pass to template

        # Products (distinct) for product filter dropdown - now from Inventory.Item
        products = cache.get("sf_products")
        if products is None:
            products = []
            try:
                qs = Item.objects.values("id", "name").order_by('name')
                for p in qs:
                    products.append({"id": p.get("id"), "name": p.get("name")})
            except Exception as e:
                print(f"⚠️ Error fetching products from Inventory: {e}")
                products = []
            cache.set("sf_products", products, timeout=self.cache_ttl)
        ctx["products"] = products

        # Recent sales (last 14 days)
        recent_sales = cache.get("sf_recent_sales")
        if recent_sales is None:
            recent_sales = []
            if get_daily_sales_df is not None:
                try:
                    end = timezone.now().date()
                    start = end - timedelta(days=30)
                    df = get_daily_sales_df(start_date=start, end_date=end)
                    # Convert pandas rows to simple dicts (date as ISO string)
                    recent_sales = [
                        {"date": r["date"].strftime("%Y-%m-%d"), "total_sales": float(r["total_quantity"])}
                        for r in df.tail(14).to_dict(orient="records")
                    ]
                except Exception:
                    recent_sales = []
            cache.set("sf_recent_sales", recent_sales, timeout=self.cache_ttl)
        ctx["actual_sales"] = recent_sales

        # Top products (last 7 days)
        top_products = cache.get("sf_top_products")
        if top_products is None:
            top_products = []
            if SaleItemUnit is not None:
                try:
                    since = timezone.now().date() - timedelta(days=7)
                    qs = (
                        SaleItemUnit.objects.filter(date__gte=since)
                        .values("product_id", "product_name")
                        .annotate(qty_sold=Sum("total_quantity"), revenue=Sum("total_revenue"))
                        .order_by("-qty_sold")[:10]
                    )
                    for r in qs:
                        top_products.append(
                            {
                                "id": r.get("product_id"),
                                "name": r.get("product_name") or f"Product {r.get('product_id')}",
                                "qty_sold": int(r.get("qty_sold") or 0),
                                "revenue": float(r.get("revenue") or 0.0),
                            }
                        )
                except Exception:
                    top_products = []
            cache.set("sf_top_products", top_products, timeout=self.cache_ttl)
        ctx["top_products"] = top_products

        # Model info + latest run summary
        latest_run = ForecastRun.objects.order_by("-created_at").first()
        if latest_run:
            model_info = {
                "name": latest_run.model_name,
                "trained_on": latest_run.train_end,
                "params": latest_run.params,
                "metrics": latest_run.metrics,
                "artifact_path": latest_run.artifact_path,
            }
            latest_run_summary = {"run_id": latest_run.id, "run_date": latest_run.created_at}
        else:
            # ✅ NEW: Use demo model info if no real model exists
            if demo_mode:
                from .demo_mode import ForecastDemoMode
                demo_info = ForecastDemoMode.generate_model_info()
                model_info = {
                    "name": demo_info['model_name'],
                    "trained_on": demo_info['trained_on'],
                    "params": demo_info['params'],
                    "metrics": demo_info['metrics'],
                    "artifact_path": None,
                }
                latest_run_summary = {
                    "run_id": 0,
                    "run_date": demo_info['last_trained'],
                    "status": "Demo Mode"
                }
            else:
                model_info = {"name": "XGBoost Regressor", "trained_on": None, "params": {}, "metrics": {}}
                latest_run_summary = None

        ctx["model_info"] = model_info
        ctx["latest_run"] = latest_run_summary

        # Small KPIs so template shows something immediately
        kpi = {"today_sales": 0.0, "forecast_sales": 0.0, "change": 0.0}
        try:
            today = timezone.now().date()
            if get_daily_sales_df is not None:
                df_today = get_daily_sales_df(start_date=today, end_date=today)
                if df_today is not None and len(df_today):
                    kpi["today_sales"] = float(df_today.iloc[-1]["total_quantity"])
            
            # ✅ NEW: Use demo KPI if in demo mode
            if demo_mode:
                from .demo_mode import ForecastDemoMode
                demo_kpi = ForecastDemoMode.generate_kpi_data()
                kpi["today_sales"] = demo_kpi['today_sales']
                kpi["forecast_sales"] = demo_kpi['next_day_forecast']
                kpi["change"] = demo_kpi['change_percent']
            else:
                if latest_run:
                    fr = ForecastResult.objects.filter(run=latest_run, product_id__isnull=True).order_by("date").first()
                    if fr:
                        kpi["forecast_sales"] = float(fr.predicted)
        except Exception:
            pass
        ctx["kpi"] = kpi

        return ctx

# ------------------------------
# Export Sales Dashboard to Excel
# ------------------------------
@login_required
def export_sales_dashboard_to_excel(request):
    """Generates and returns an Excel file of the current sales dashboard data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Dashboard Report"

    # --- Section 1: Top Products (last 7 days) ---
    ws.append(["Top Products (Last 7 Days)"])
    ws.append(['Product Name', 'Quantity Sold', 'Revenue'])
    try:
        seven_days_ago = timezone.now().date() - timedelta(days=7)
        top_products_qs = (
            SaleItemUnit.objects.filter(date__gte=seven_days_ago)
            .values('product_name')
            .annotate(qty_sold=Sum('total_quantity'), revenue=Sum('total_revenue'))
            .order_by('-qty_sold')[:10]
        )
        for p in top_products_qs:
            ws.append([p['product_name'], p['qty_sold'], float(p['revenue'])])
    except Exception as e:
        ws.append([f"Error fetching top products: {e}"])
    ws.append([]) # Blank row for separation

    # --- Section 2: Monthly Sales --- 
    ws.append(["Monthly Sales (Last 13 Months)"])
    ws.append(['Month', 'Total Sales'])
    try:
        monthly_sales_data = (
            DailySalesRecord.objects
            .annotate(month=TruncMonth('date'))
            .values('month')
            .annotate(total_sales=Sum('total_sales'))
            .order_by('month')
        )


        for ms in monthly_sales_data:
            ws.append([ms['month'].strftime('%Y-%m'), float(ms['total_sales'])])
    except Exception as e:
        ws.append([f"Error fetching monthly sales: {e}"])
    ws.append([]) # Blank row for separation

    # --- Section 3: Products to Restock (based on next prediction date) ---
    ws.append(["Products to Restock (Next Predicted Sale Date)"])
    ws.append(['Product Name', 'Prediction Date', 'Predicted Sales Count', 'Current Stock', 'Min Stock Level'])
    try:
        # Get the latest forecast run
        latest_run = ForecastRun.objects.order_by('-created_at').first()
        if latest_run:
            # Filter for product-specific forecasts, ordered by date
            restock_predictions = ForecastResult.objects.filter(
                run=latest_run, 
                product__isnull=False # Only product-specific predictions
            ).order_by('date', 'product__name')

            seen_products = set()
            for fr in restock_predictions:
                # Use fr.product.name directly and check if predicted quantity is below min_stock_level for restock alert
                if fr.product and fr.product.id not in seen_products:
                    # Heuristic: if predicted sales exceed current stock or fall below min_stock_level
                    # For simplicity, let's recommend restock if predicted units are > 0 and stock is below min_stock_level.
                    # A more complex heuristic would involve comparing predicted units with current stock and min_stock_level.
                    current_stock = fr.product.stock if fr.product else 0
                    min_stock_level = fr.product.min_stock_level if fr.product else 0

                    # If current stock + predicted units for next day < min_stock_level, suggest restock
                    if current_stock < min_stock_level and fr.predicted > 0:
                         ws.append([fr.product.name, fr.date.strftime('%Y-%m-%d'), round(fr.predicted), current_stock, min_stock_level])
                         seen_products.add(fr.product.id)

        else:
             ws.append(["No forecast runs available for restock predictions."])
    except Exception as e:
        ws.append([f"Error fetching restock predictions: {e}"])
    ws.append([]) # Blank row for separation

    # Prepare the Excel file for download
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    current_month_year = timezone.now().strftime('%Y-%m')
    filename = f"sales_dashboard_report_{current_month_year}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response












# ==================== FORECAST REPORT VIEW ====================
@login_required
def forecast_report_view(request):
    download_excel = request.GET.get('excel', 'false').lower() == 'true'

    # Get the latest forecast run
    latest_run = ForecastRun.objects.order_by('-created_at').first()

    forecast_data = []
    if latest_run:
        # Group forecast results by product and sum predicted units for each date
        # This assumes 'predicted' represents units/quantity for restock prediction
        forecast_results = ForecastResult.objects.filter(run=latest_run).order_by('date', 'product__name')

        for result in forecast_results:
            product_name = result.product.name if result.product else "Total Sales"
            forecast_data.append({
                'date': result.date,
                'product_name': product_name,
                'predicted_units': round(result.predicted),
                'total_revenue': result.predicted * result.product.price if result.product and result.product.price else None # Assuming predicted units are needed for total revenue estimation
            })

    if download_excel:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Forecast Report"

        # Add header row
        header = ['Date', 'Product to Restock', 'Units', 'Estimated Revenue']
        sheet.append(header)

        # Style header
        header_font = Font(bold=True)
        for col_idx in range(1, len(header) + 1):
            sheet.cell(row=1, column=col_idx).font = header_font

        # Add data rows
        for data in forecast_data:
            sheet.append([
                data['date'].isoformat(),
                data['product_name'],
                data['predicted_units'],
                float(data['total_revenue']) if data['total_revenue'] is not None else ''
            ])

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="forecast_report.xlsx"'
        workbook.save(response)
        return response

    context = {
        'latest_run': latest_run,
        'forecast_data': forecast_data,
    }
    return render(request, 'Sales_forecast/forecast_report.html', context)